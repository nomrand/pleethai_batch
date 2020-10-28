import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color

# Excel Index Color
COLOR_INDEX = (
    '00FF0000', '0000FF00', '000000FF',  # 0-4
    '00FFFF00', '00FF00FF', '0000FFFF',  # 5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF',  # 10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000',  # 15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF',  # 20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080',  # 25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00',  # 30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF',  # 35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF',  # 40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC',  # 45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699',  # 50-54
    '00969696', '00003366', '00339966', '00003300', '00333300',  # 55-59
    '00993300', '00993366', '00333399', '00333333',  # 60-63
)


class WS():
    def __init__(self, filename, sheetname, needsave=False):
        self.filename = filename
        self.sheetname = sheetname
        self.needsave = needsave
        self.wb = None

    def __enter__(self):
        # open workbook
        self.wb = openpyxl.load_workbook(self.filename)
        # get worksheet
        if self.sheetname == '' or self.sheetname is None:
            # first sheet
            ws = self.wb.worksheets[0]
        else:
            ws = self.wb[self.sheetname]
        return ws

    def __exit__(self, exception_type, exception_value, traceback):
        if self.needsave:
            self.wb.save(self.filename)


def get_excel_data(filename, sheetname=None):
    ''' Get Data from Excel Sheet
    The 1st row must be header. Get only rows with data. 

    Parameters
    ----------
    filename : str
        File path to Excel
    sheetname : str
        Sheet name to access (default is the 1st sheet)

    Returns
    -------
    result : dict
        key:str of header(1st row)
        value:str
    '''
    lines = []
    with WS(filename, sheetname) as ws:
        # convert excel to dict
        for row_i, row in enumerate(ws.rows):
            # first row
            if row_i == 0:
                first = [col.value for col in ws.rows.__next__()]
                continue
            # skip all None line
            if all(c.value is None for c in row):
                continue

            # second row to end
            line = {}
            for i, cell in enumerate(row):
                line[first[i]] = cell.value
            lines.append(line)
    return lines


def set_excel_data(lines, filename, sheetname=None, startrow=0, color=None):
    ''' Set Data to Excel Sheet

    Parameters
    ----------
    filename : str
        File path to Excel
    sheetname : str
        Sheet name to access (default is the 1st sheet)
    startrow : int
        Ther row number to write data (1-based index) (default is 0)
        If startrow <= 1 (1st row must be header), write data to empty rows
    color : str
        Cell style color code for new data (such as '00112233')
    '''
    with WS(filename, sheetname, True) as ws:
        rows = list(ws.rows)
        # first row
        first = [col.value for col in rows[0]]

        indexrow = max(startrow, 2)
        for i, line in enumerate(lines):
            if startrow <= 1:
                # skip non None line
                while indexrow+i-1 < len(rows) and any(col.value is not None for col in rows[indexrow+i-1]):
                    indexrow += 1

            for key in line.keys():
                # find the column to set from first row
                # (key is the header title)
                if key not in first:
                    continue
                cell = ws.cell(row=indexrow + i,
                               column=first.index(key) + 1,
                               value=line[key])
                if color is not None:
                    cell = ws.cell(row=indexrow + i, column=1)
                    cell.fill = PatternFill(
                        patternType='solid', fgColor=color)


if __name__ == '__main__':
    import sys
    import random

    if (len(sys.argv) < 2) or (len(sys.argv) > 4):
        print('ERROR')
        print('Usage: # python %s excel_filename [read_sheet_name [write_sheet_name]]' %
              sys.argv[0])
        quit()

    filename = sys.argv[1]
    rd_sheetname = None
    wr_sheetname = None
    if len(sys.argv) >= 3:
        rd_sheetname = sys.argv[2]
    if len(sys.argv) == 4:
        wr_sheetname = sys.argv[3]

    res = get_excel_data(filename, rd_sheetname)
    set_excel_data(res, filename, wr_sheetname,
                   color=random.choice(COLOR_INDEX))
